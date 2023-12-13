# ----------------- STAIRCASE
class Staircase:
    """
    A staircase procedure implimenting the Kesten algorithm.
    Allows target threshold to be specified directly,
    e.g., 0.5 for PSE, 0.75 for a typical threshold,
    or any value you want. It is uses an efficient algorithm
    that allows for fewer trials than usual (15 trials is 
    often sufficient to converge onto desired threshold).     
    Staircase ends when minimum reversals and trials reached.
    
    Arguments:
        name: string
        type: 'simple' or 'accel'.
              'simple': step sizes change after each reversal
              'accel': step sizes change after each trial for 
                       the first 2 trials. Thereafter, change after
                       each reversal.
        value: numeric. Initial staircase value
        C: numeric. Arbitrary value, typically 60% of reference stimulus.
                    The higher the number, the larger the initial stepsize.
        minRevs: numeric. Minimum number of reversals for completion.
        minTrials: numeric. Minumum number of trials for completion.
        targetThresh: numeric. Any threshold, e.g., 0.5 for PSE,
                               0.75 for typical threshold. Even a crazy 
                               0.87 for whatever reason.
        minVal: numeric. Minumum value for staircase, default is 0. 
        
    Example use:
    Consider a temporal discrimination task with 2 intervals.
    The first interval is a fixed duration of 60 frames, the 
    second interval is staircase controlled. Initial value is
    set to 90 frames (should be easy to detect). The 'C' argument
    is arbitrary but usually 30-60% of fixed simulus (e.g., 
    60% of 60 is 36, in this example).
    myStaircase = Staircase(name='myKesten', 
                            type='simple',
                            value=90, 
                            C=36, # 60 % of reference (60 frames) duration
                            minRevs=4, # finish when minRevs 
                            minTrials=15, # and minTrials achieved
                            targetThresh=0.5, # aim for PSE
                            minVal=1)
    Use:
        for step in myStaircase:
            # draw stimuli 
            intensity = step
            jokingResponse = random.choice([0,1])
            resp = myStaircase.newValue(jokingResponse)
        
        or
        
        while something is True:            
            # draw stimuli
            intensity = myStaircase.next()
            jokingResponse = random.choice([0,1])
            resp = myStaircase.newValue(jokingResponse)            
      
    """
  
    def __init__(self, name=' ',type='simple',value=100,C=60,minRevs=4,minTrials=15,targetThresh=0.5,minVal=0,maxVal=1, extraInfo=None):
        self.name = name
        self.value = value    
        self.respList = []     
        self.valueList = []        
        self.C = C
        self.trialCount = -1
        self.reversals=[]
        rList=[]
        self.minRevs=minRevs
        self.minTrials=minTrials-1
        self.targetThresh = targetThresh
        self.minVal = minVal
        self.maxVal = maxVal
        self.type=type
        self.extraInfo = extraInfo
        
    def __iter__(self):
        return self
    
    def __str__(self):
        return self.name
    
    def countReversals(self): # could probably be a private function.
        self.reversals=[]
        for index,item in enumerate(self.respList):
            if self.respList[index-1] != self.respList[index]:
                if index != 0:
                    self.reversals.append(index)      
        return len(self.reversals)
        
    def newValue(self,resp): # calculates value for next trial
        self.resp = resp          
        self.respList.append(resp)
        if self.type == 'simple':
            formula = (self.C*(self.resp - self.targetThresh) / (1 + self.countReversals()))
        elif self.type == 'accel':
            if len(self.respList) < 3:
                formula = (self.C*(self.resp - self.targetThresh) / (1 + self.trialCount))
            elif len(self.respList) > 2:
                formula = (self.C*(self.resp - self.targetThresh) / (2 + self.countReversals()))
        newFigure = self.value - formula         
        self.value=newFigure
        if self.value < self.minVal:
            self.value = self.minVal
        if self.value > self.maxVal:
            self.value = self.maxVal
          

                  
    def reversalIndices(self): # also could be private function
        return self.reversals

    def getReversals(self): # returns list of reversal intensities
        rList=[]
        for i in self.reversalIndices():            
            v = self.valueList[i]
            rList.append(v)
        return rList
    
    def getValues(self): # returns list of intensities
        return self.valueList

    def getResps(self): # returns list of responses
        return self.respList
        
    def printValue(self): # returns current staircase intensity 
        return self.value # same as next()
            
    def next(self):       
        if self.countReversals() >= self.minRevs and self.trialCount >= self.minTrials:
            raise StopIteration
        else:
            self.trialCount += 1
            self.valueList.append(self.value)
            return self.value
    
    def saveAsExcel(self, fileName, sheetName='data',
                        matrixOnly=False, appendFile=True,
                        fileCollisionMethod='rename'):
            """Save a summary data file in Excel OpenXML format workbook
            (:term:`xlsx`) for processing in most spreadsheet packages.
            This format is compatible with versions of Excel (2007 or greater)
            and and with OpenOffice (>=3.0).
    
            It has the advantage over the simpler text files
            (see :func:`TrialHandler.saveAsText()` ) that data can be stored
            in multiple named sheets within the file. So you could have a
            single file named after your experiment and then have one worksheet
            for each participant. Or you could have one file for each participant
            and then multiple sheets for repeated sessions etc.
    
            The file extension `.xlsx` will be added if not given already.
    
            The file will contain a set of values specifying the staircase level
            ('intensity') at each reversal, a list of reversal indices
            (trial numbers), the raw staircase / intensity level on *every*
            trial and the corresponding responses of the participant on every
            trial.
    
            :Parameters:
    
                fileName: string
                    the name of the file to create or append. Can include
                    relative or absolute path.
    
                sheetName: string
                    the name of the worksheet within the file
    
                matrixOnly: True or False
                    If set to True then only the data itself will be output
                    (no additional info)
    
                appendFile: True or False
                    If False any existing file with this name will be
                    overwritten. If True then a new worksheet will be appended.
                    If a worksheet already exists with that name a number will
                    be added to make it unique.
    
                fileCollisionMethod: string
                    Collision method passed to
                    :func:`~psychopy.tools.fileerrortools.handleFileCollision`
                    This is ignored if ``appendFile`` is ``True``.
    
            """
    
            if self.thisTrialN < 1:
                if self.autoLog:
                    logging.debug('StairHandler.saveAsExcel called but no '
                                  'trials completed. Nothing saved')
                return -1
            # NB this was based on the limited documentation for openpyxl v1.0
            if not haveOpenpyxl:
                raise ImportError('openpyxl is required for saving files in '
                                  'Excel (xlsx) format, but was not found.')
                # return -1
    
            # import necessary subpackages - they are small so won't matter to do
            # it here
            from openpyxl.workbook import Workbook
            from openpyxl.reader.excel import load_workbook
    
            if not fileName.endswith('.xlsx'):
                fileName += '.xlsx'
            # create or load the file
            if appendFile and os.path.isfile(fileName):
                wb = load_workbook(fileName)
                newWorkbook = False
            else:
                if not appendFile:
                    # the file exists but we're not appending, will be overwritten
                    fileName = handleFileCollision(fileName,
                                                   fileCollisionMethod)
                wb = Workbook()
                wb.properties.creator = 'PsychoPy' + psychopy.__version__
                newWorkbook = True
    
            if newWorkbook:
                ws = wb.worksheets[0]
                ws.title = sheetName
            else:
                ws = wb.create_sheet()
                ws.title = sheetName
    
            # write the data
            # reversals data
            ws['A1'] = 'Reversal Intensities'
            ws['B1'] = 'Reversal Indices'
            for revN, revIntens in enumerate(self.reversalIntensities):
                ws.cell(column=1, row=revN+2,
                        value=u"{}".format(revIntens))
                ws.cell(column=2, row=revN+2,
                        value=u"{}".format(self.reversalPoints[revN]))
    
            # trials data
            ws['C1'] = 'All Intensities'
            ws['D1'] = 'All Responses'
            for intenN, intensity in enumerate(self.intensities):
                ws.cell(column=3, row=intenN+2,
                        value=u"{}".format(intensity))
                ws.cell(column=4, row=intenN+2,
                        value=u"{}".format(self.data[intenN]))
    
            # add other data
            col = 5
            if self.otherData is not None:
                # for varName in self.otherData:
                for key, val in list(self.otherData.items()):
                    ws.cell(column=col, row=1,
                            value=u"{}".format(key))
                    for oDatN in range(len(self.otherData[key])):
                        ws.cell(column=col, row=oDatN+2,
                                value=u"{}".format(self.otherData[key][oDatN]))
                    col += 1
    
            # add self.extraInfo
            if self.extraInfo is not None and not matrixOnly:
                ws.cell(column=startingCol, row=1,
                        value='extraInfo')
                rowN = 2
                for key, val in list(self.extraInfo.items()):
                    ws.cell(column=col, row=rowN,
                            value=u"{}:".format(key))
                    _cell = _getExcelCellName(col=col+1, row=rowN)
                    ws.cell(column=col+2, row=rowN+1,
                            value=u"{}".format(val))
                    rowN += 1
    
    
            wb.save(filename=fileName)
            if self.autoLog:
                logging.info('saved data to %s' % fileName)
